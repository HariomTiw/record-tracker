import os
import re
from datetime import datetime
from openpyxl import load_workbook
from gemini_extractor import extract_meal_data
from difflib import get_close_matches


def normalize_name(name):
    if not name:
        return ""

    # Remove extra spaces
    name = re.sub(r"\s+", " ", str(name))

    # Remove invisible unicode junk
    name = name.replace("\u200b", "")

    return name.strip()

def clean_records(data):

    cleaned = []

    for record in data.get("records", []):

        try:
            sr_no = int(str(record.get("sr_no", 0)).strip())

            name = str(record.get("name", "")).strip()

            morning = int(record.get("morning", 0))
            afternoon = int(record.get("afternoon", 0))
            night = int(record.get("night", 0))

            cleaned.append({
                "sr_no": sr_no,
                "name": name,
                "morning": 1 if morning == 1 else 0,
                "afternoon": 1 if afternoon == 1 else 0,
                "night": 1 if night == 1 else 0
            })

        except Exception:
            print("Skipping corrupted record:", record)

    return {"records": cleaned}


def save_to_excel(data):

    date_obj = datetime.strptime(data["date"], "%Y-%m-%d")
    month_name = date_obj.strftime("%b").upper()
    year_short = date_obj.strftime("%y")
    day = date_obj.day

    filename = f"{month_name}-{year_short}.xlsx"

    if not os.path.exists(filename):
        print(f"{filename} not found.")
        return

    wb = load_workbook(filename)
    ws = wb.active

    # Column index for that day
    # Column C = day 1 → index 3
    day_column = day + 2

    # Create lookup dictionary
    excel_lookup = {}

    for row in ws.iter_rows(min_row=2):
        excel_name = normalize_name(row[0].value)
        session = normalize_name(row[1].value)
        if excel_name:
         excel_lookup.setdefault(excel_name, {})[session] = row[0].row
    # Now update
    for record in data["records"]:
        name = normalize_name(record["name"])

        matched_name = None

        if name in excel_lookup:
            matched_name = name
        else:
            # Try fuzzy match
            matches = get_close_matches(name, excel_lookup.keys(), n=1, cutoff=0.90)
            if matches:
                matched_name = matches[0]
                print(f"Fuzzy matched: {name} → {matched_name}")

        if matched_name:

            if "સવાર" in excel_lookup[matched_name]:
                ws.cell(
                    row=excel_lookup[matched_name]["સવાર"],
                    column=day_column
                ).value = record["morning"]

            if "બપોર" in excel_lookup[matched_name]:
                ws.cell(
                    row=excel_lookup[matched_name]["બપોર"],
                    column=day_column
                ).value = record["afternoon"]

            if "રાત્રે" in excel_lookup[matched_name]:
                ws.cell(
                    row=excel_lookup[matched_name]["રાત્રે"],
                    column=day_column
                ).value = record["night"]
        else:
            print("Still not found:", name)

    wb.save(filename)
    print(f"{filename} updated successfully.")
if __name__ == "__main__":

    while True:
        user_date = input("Enter date (YYYY-MM-DD): ").strip()
        try:
            datetime.strptime(user_date, "%Y-%m-%d")
            break
        except ValueError:
            print("Invalid format. Please enter in YYYY-MM-DD format.")



    data = extract_meal_data("sample.jpg")
    data = clean_records(data)
    print("Total extracted:", len(data["records"]))

    # Overwrite Gemini date with user date
    data["date"] = user_date

    save_to_excel(data)

    print("Excel updated successfully!")
    print(data)